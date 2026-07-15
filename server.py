"""
Backend for the Excel to XLPivot & Reporting System page.

Each report button on the HTML page sends a POST request to /api/run-report
with:
  - file:       the Excel/CSV file picked by the user
  - report_id:  a string identifying which button was clicked

This server routes each report_id to its own Python function below.
Replace the body of each function with your converted VBA logic.

LOCAL USE (just you, on your own computer):
    pip install -r requirements.txt
    python server.py
    -> open http://127.0.0.1:5000 in the browser

DEPLOYING SO OTHERS CAN USE IT (see the deployment notes at the bottom
of this file / the accompanying instructions) runs this instead with:
    gunicorn server:app
"""

from flask import Flask, request, jsonify, send_file, send_from_directory
import os
import uuid
import tempfile
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Serves login_form.html / dashboard.html directly from this same app,
# so the whole thing is one deployable unit (one URL, no separate hosting
# for the HTML files, and no cross-origin/CORS headaches).
app = Flask(__name__, static_folder=".", static_url_path="")

BASE_TMP_DIR = tempfile.mkdtemp(prefix="xlpivot_")


# ---------------------------------------------------------------------------
# One function per report button.
# Each takes the path to the uploaded source file and must return EITHER:
#   - a dict, e.g. {"message": "Done, saved to X"}   -> shown as a status message
#   - a file path (string) to an Excel/PDF/CSV file  -> sent back for download
# ---------------------------------------------------------------------------


def now_ist():
    """India doesn't observe DST, so a fixed UTC+5:30 offset is always
    correct -- this avoids depending on system/tzdata availability on
    whatever server this app is deployed to."""
    return datetime.utcnow() + timedelta(hours=5, minutes=30)


def match_columns_case_insensitive(df, required_names):
    """
    Finds each required column name in df regardless of case/extra spaces
    (e.g. 'Business unit' vs 'Business Unit'), and returns a mapping of
    {requested_name: actual_column_name_in_df}. Raises ValueError listing
    anything genuinely missing.
    """
    lookup = {str(c).strip().lower(): c for c in df.columns}
    mapping = {}
    missing = []
    for name in required_names:
        key = name.strip().lower()
        if key in lookup:
            mapping[name] = lookup[key]
        else:
            missing.append(name)
    if missing:
        raise ValueError("Missing expected column(s): " + ", ".join(missing))
    return mapping


def dedupe_headers(headers):
    """
    Some source files repeat the exact same header text for more than one
    column (Excel's PivotTable engine silently renames the 2nd/3rd
    occurrence to 'Header2'/'Header3' internally for its own field list --
    the cells themselves still say the same thing). This reproduces that
    same renaming so repeated columns become individually addressable.
    """
    seen = {}
    result = []
    for h in headers:
        key = "" if h is None else str(h).strip()
        if key == "":
            result.append(h)
            continue
        if key not in seen:
            seen[key] = 1
            result.append(key)
        else:
            seen[key] += 1
            result.append(f"{key}{seen[key]}")
    return result


def write_hierarchical_table(ws_out, df, group_levels, all_columns, value_cols,
                              start_row, blank_line_levels=None,
                              header_fill_color="BDD7EE",
                              subtotal_fill_color="DDEBF7",
                              grand_total_fill_color="FFEB9C"):
    """
    Writes df as a detail table with a subtotal row inserted after every
    group in `group_levels` (outermost first), the same way Excel's
    PivotTable/Subtotal features nest multiple row fields that each have
    their own subtotal turned on. Ends with one Grand Total row.

    - all_columns: full ordered list of columns to print per detail row
      (including the group_levels columns and any extra detail-only columns)
    - value_cols: the numeric columns to sum for subtotal/grand total rows
    - blank_line_levels: set of level names that get a blank row inserted
      right after their subtotal row (mirrors VBA's LayoutBlankLine)

    Returns the next free row after everything has been written.
    """
    blank_line_levels = blank_line_levels or set()
    row = {"i": start_row}

    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    subtotal_fill = PatternFill(start_color=subtotal_fill_color, end_color=subtotal_fill_color, fill_type="solid")
    grand_fill = PatternFill(start_color=grand_total_fill_color, end_color=grand_total_fill_color, fill_type="solid")
    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for j, label in enumerate(all_columns, start=1):
        cell = ws_out.cell(row=row["i"], column=j, value=label)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border
    row["i"] += 1

    def recurse(sub_df, levels):
        if not levels:
            for _, rec in sub_df.iterrows():
                for j, col in enumerate(all_columns, start=1):
                    ws_out.cell(row=row["i"], column=j, value=rec[col]).border = thin_border
                row["i"] += 1
            return {c: float(sub_df[c].sum()) for c in value_cols}

        level = levels[0]
        level_totals = {c: 0.0 for c in value_cols}
        for key, group in sub_df.groupby(level, sort=False):
            sub_totals = recurse(group, levels[1:])

            label_col_idx = all_columns.index(level) + 1
            for col_idx in range(1, len(all_columns) + 1):
                cell = ws_out.cell(row=row["i"], column=col_idx)
                cell.fill = subtotal_fill
                cell.border = thin_border
                if col_idx == label_col_idx:
                    cell.value = f"{key} Total"
                    cell.font = Font(bold=True)
            for c in value_cols:
                col_idx = all_columns.index(c) + 1
                cell = ws_out.cell(row=row["i"], column=col_idx, value=sub_totals[c])
                cell.font = Font(bold=True)
                cell.fill = subtotal_fill
                cell.border = thin_border
            row["i"] += 1

            if level in blank_line_levels:
                row["i"] += 1

            for c in value_cols:
                level_totals[c] += sub_totals[c]
        return level_totals

    grand_totals = recurse(df, group_levels)

    grand_row = row["i"]
    for col_idx in range(1, len(all_columns) + 1):
        cell = ws_out.cell(row=grand_row, column=col_idx)
        cell.fill = grand_fill
        cell.border = thin_border
        if col_idx == 1:
            cell.value = "Grand Total"
            cell.font = Font(bold=True)
    for c in value_cols:
        col_idx = all_columns.index(c) + 1
        cell = ws_out.cell(row=grand_row, column=col_idx, value=grand_totals[c])
        cell.font = Font(bold=True)
        cell.fill = grand_fill
        cell.border = thin_border
    row["i"] += 1

    return row["i"]


def find_header_row_auto(ws, max_scan=20):
    """Equivalent of VBA FindHeaderRow: detects the header row by scanning
    the first few rows for known label text, or the first row where the
    first three columns are all non-empty."""
    for r in range(1, max_scan + 1):
        v = str(ws.cell(row=r, column=1).value or "").strip().lower()
        if v in ("sl.no.", "sl no", "sl.no", "employee code"):
            return r
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        if a not in (None, "") and b not in (None, "") and c not in (None, ""):
            return r
    return 0


def build_salary_pivot_table(input_path):
    """
    Shared computation used by both salary_pivot_bu_wise (Excel output) and
    salary_pivot_bu_wise_pdf (PDF output), converted from VBA:
    CreatePivotFromSourceData.

    Returns (detail_df, summary_df, group_cols, value_cols).
    """
    wb = load_workbook(input_path, data_only=True)
    ws = wb.worksheets[0]

    header_row = find_header_row_auto(ws)
    if header_row == 0:
        raise ValueError("Could not detect header row in SourceData.")

    last_col = 1
    for c in range(ws.max_column, 0, -1):
        if ws.cell(row=header_row, column=c).value not in (None, ""):
            last_col = c
            break

    last_row = header_row
    for r in range(ws.max_row, header_row, -1):
        if ws.cell(row=r, column=1).value not in (None, ""):
            last_row = r
            break

    if last_row <= header_row:
        raise ValueError("No data rows found below the header in SourceData.")

    headers = [ws.cell(row=header_row, column=c).value for c in range(1, last_col + 1)]
    rows = [[ws.cell(row=r, column=c).value for c in range(1, last_col + 1)]
            for r in range(header_row + 1, last_row + 1)]
    df = pd.DataFrame(rows, columns=headers)

    group_cols = ["SalaryBU", "Department", "Employee Name", "Designation"]
    value_cols_map = {
        "Gross Salary": "Gross Salary",
        "[2_DED Income Tax]": "DED Income Tax",
        "[2_DED Advance]": "DED Advance",
        "[2_DED PF]": "DED PF",
        "Deduct Salary": "Deduct Salary",
        "Net Salary": "Net Salary",
    }

    # Robust match (case/whitespace-insensitive) so small header differences
    # in the real file don't break this.
    col_map = match_columns_case_insensitive(df, group_cols + list(value_cols_map.keys()))
    df = df.rename(columns={actual: wanted for wanted, actual in col_map.items()})

    for src_col in value_cols_map:
        df[src_col] = pd.to_numeric(df[src_col], errors="coerce").fillna(0)

    df = df.sort_values(by=group_cols).reset_index(drop=True)
    df = df.rename(columns=value_cols_map)
    value_cols = list(value_cols_map.values())

    detail_df = df[group_cols + value_cols].copy()
    summary_df = (
        df.groupby("SalaryBU", dropna=False)[["Gross Salary", "Deduct Salary", "Net Salary"]]
        .sum()
        .reset_index()
    )

    return detail_df, summary_df, group_cols, value_cols


def salary_pivot_bu_wise(input_path):
    """
    Converted from VBA: CreatePivotFromSourceData (CommandButton6, behind
    "1. Salary Pivot BU Wise").

    Reproduces the SalaryBU > Department > Employee Name > Designation
    grouping/sort order, WITH a subtotal row for both SalaryBU and
    Department (matching the VBA, which had subtotals turned on for both of
    those fields), plus the bottom per-SalaryBU summary block. Excel's
    native collapsible PivotTable grouping UI itself can't be reproduced
    outside Excel -- this gives the same totals/rows instead.
    """
    detail_df, summary_df, group_cols, value_cols = build_salary_pivot_table(input_path)

    wb = load_workbook(input_path)
    new_sheet_name = "Pivot_" + now_ist().strftime("%H%M%S")
    ws_out = wb.create_sheet(new_sheet_name)

    header_row_out = 10  # matches TableDestination A10 in the VBA
    all_columns = group_cols + value_cols

    next_row = write_hierarchical_table(
        ws_out, detail_df,
        group_levels=["SalaryBU", "Department"],
        all_columns=all_columns,
        value_cols=value_cols,
        start_row=header_row_out,
        blank_line_levels={"SalaryBU", "Department"},
    )

    row_idx = next_row + 2

    # Bottom summary block (per SalaryBU: Gross / Deduct / Net) -- mirrors AddBottomSummary
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    for j, label in enumerate(["SalaryBU", "Gross Salary", "Deduct Salary", "Net Salary"], start=1):
        cell = ws_out.cell(row=row_idx, column=j, value=label)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    row_idx += 1

    g_gross = g_deduct = g_net = 0.0
    for _, rec in summary_df.iterrows():
        ws_out.cell(row=row_idx, column=1, value=f"{rec['SalaryBU']} Total")
        ws_out.cell(row=row_idx, column=2, value=float(rec["Gross Salary"]))
        ws_out.cell(row=row_idx, column=3, value=float(rec["Deduct Salary"]))
        ws_out.cell(row=row_idx, column=4, value=float(rec["Net Salary"]))
        g_gross += float(rec["Gross Salary"])
        g_deduct += float(rec["Deduct Salary"])
        g_net += float(rec["Net Salary"])
        row_idx += 1

    grand_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    for j, val in enumerate(["Grand Total", g_gross, g_deduct, g_net], start=1):
        cell = ws_out.cell(row=row_idx, column=j, value=val)
        cell.font = Font(bold=True)
        cell.fill = grand_fill

    for col_idx in range(1, len(all_columns) + 1):
        ws_out.column_dimensions[get_column_letter(col_idx)].width = 18

    wb.save(input_path)
    return input_path


def salary_pivot_bu_wise_pdf(input_path):
    """
    Converted from VBA: CommandButton5_Click (exports the pivot to PDF).
    The VBA version exported whichever pivot was created most recently in
    the same session; this web version just rebuilds the same table and
    exports it straight to PDF -- including the SalaryBU AND Department
    subtotal rows (colored so they stand out), matching the Excel version.
    """
    detail_df, summary_df, group_cols, value_cols = build_salary_pivot_table(input_path)

    pdf_path = os.path.splitext(input_path)[0] + "_Salary_Pivot_BU_Wise.pdf"
    doc = SimpleDocTemplate(
        pdf_path, pagesize=landscape(A4),
        leftMargin=0.3 * inch, rightMargin=0.3 * inch,
        topMargin=0.3 * inch, bottomMargin=0.3 * inch,
    )
    styles = getSampleStyleSheet()
    elements = [Paragraph("Salary Pivot BU Wise", styles["Title"]), Spacer(1, 12)]

    header = group_cols + value_cols
    data = [header]
    subtotal_rows = []   # track which data rows are subtotal rows, for styling
    grand_total_row_idx = None

    grand_totals = {c: 0.0 for c in value_cols}

    for bu_val, bu_group in detail_df.groupby("SalaryBU", sort=False):
        bu_totals = {c: 0.0 for c in value_cols}

        for dept_val, dept_group in bu_group.groupby("Department", sort=False):
            for _, rec in dept_group.iterrows():
                data.append([rec[c] for c in header])
            dept_totals = [float(dept_group[c].sum()) for c in value_cols]
            data.append([f"{dept_val} Total", "", "", ""] + [round(v, 2) for v in dept_totals])
            subtotal_rows.append(len(data) - 1)
            for c, v in zip(value_cols, dept_totals):
                bu_totals[c] += v

        data.append([f"{bu_val} Total", "", "", ""] + [round(bu_totals[c], 2) for c in value_cols])
        subtotal_rows.append(len(data) - 1)
        for c in value_cols:
            grand_totals[c] += bu_totals[c]

    data.append(["Grand Total", "", "", ""] + [round(grand_totals[c], 2) for c in value_cols])
    grand_total_row_idx = len(data) - 1

    table = Table(data, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#BDD7EE")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
    ]
    for r in subtotal_rows:
        style_cmds.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#DDEBF7")))
        style_cmds.append(("FONTNAME", (0, r), (-1, r), "Helvetica-Bold"))
    style_cmds.append(("BACKGROUND", (0, grand_total_row_idx), (-1, grand_total_row_idx), colors.HexColor("#FFEB9C")))
    style_cmds.append(("FONTNAME", (0, grand_total_row_idx), (-1, grand_total_row_idx), "Helvetica-Bold"))

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)
    doc.build(elements)
    return pdf_path


def _compare_salary_field(input_path, input_path2, field_name):
    """
    Shared logic for gross_salary_comparision / net_salary_comparision
    (converted from VBA: CompareGrossSalary / CompareNetSalary -- they were
    identical except for which salary column is compared).
    """
    HEADER_ROW = 10  # headers on row 10, data starts row 11

    def load_sheet(path):
        wb = load_workbook(path, data_only=True)
        if "SourceData" not in wb.sheetnames:
            raise ValueError(f"Sheet 'SourceData' not found in {os.path.basename(path)}")
        return wb, wb["SourceData"]

    wb1, ws1 = load_sheet(input_path)
    wb2, ws2 = load_sheet(input_path2)

    needed = ["Employee Code", "SalaryBU", "Department", "Employee Name", "Designation", field_name]

    def read_records(ws):
        last_col = 1
        for c in range(ws.max_column, 0, -1):
            if ws.cell(row=HEADER_ROW, column=c).value not in (None, ""):
                last_col = c
                break
        cols = {name: find_header_column(ws, HEADER_ROW, last_col, name) for name in needed}
        missing = [n for n, c in cols.items() if c == 0]
        if missing:
            raise ValueError(f"Required header(s) {', '.join(missing)} not found in row {HEADER_ROW}.")

        last_row = ws.max_row
        for r in range(ws.max_row, HEADER_ROW, -1):
            if ws.cell(row=r, column=cols["Employee Code"]).value not in (None, ""):
                last_row = r
                break

        records = {}
        for r in range(HEADER_ROW + 1, last_row + 1):
            code = str(ws.cell(row=r, column=cols["Employee Code"]).value or "").strip()
            if code == "" or code in records:
                continue
            val = ws.cell(row=r, column=cols[field_name]).value
            records[code] = {
                "SalaryBU": ws.cell(row=r, column=cols["SalaryBU"]).value,
                "Department": ws.cell(row=r, column=cols["Department"]).value,
                "Employee Name": ws.cell(row=r, column=cols["Employee Name"]).value,
                "Designation": ws.cell(row=r, column=cols["Designation"]).value,
                "value": float(val) if isinstance(val, (int, float)) else 0.0,
            }
        return records

    dict1 = read_records(ws1)
    dict2 = read_records(ws2)

    all_codes = list(dict1.keys()) + [k for k in dict2 if k not in dict1]

    n = 1
    while str(n) in wb2.sheetnames:
        n += 1
    result_ws = wb2.create_sheet(str(n))

    headers = ["Employee Code", "SalaryBU", "Department", "Employee Name", "Designation",
               f"{field_name} (File 1)", f"{field_name} (File 2)", "Difference", "Status"]
    for j, h in enumerate(headers, start=1):
        c = result_ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="C8DCFF", end_color="C8DCFF", fill_type="solid")

    row_idx = 2
    for code in all_codes:
        found1 = code in dict1
        found2 = code in dict2
        rec1 = dict1.get(code)
        rec2 = dict2.get(code)
        base = rec1 if found1 else rec2

        g1 = rec1["value"] if found1 else 0.0
        g2 = rec2["value"] if found2 else 0.0
        diff = g1 - g2

        if found1 and not found2:
            status = "Deleted in File 2"
        elif found2 and not found1:
            status = "New in File 2"
        elif diff != 0:
            status = "Changed"
        else:
            status = "Same"

        result_ws.cell(row=row_idx, column=1, value=code)
        result_ws.cell(row=row_idx, column=2, value=base["SalaryBU"])
        result_ws.cell(row=row_idx, column=3, value=base["Department"])
        result_ws.cell(row=row_idx, column=4, value=base["Employee Name"])
        result_ws.cell(row=row_idx, column=5, value=base["Designation"])
        result_ws.cell(row=row_idx, column=6, value=g1 if found1 else None)
        result_ws.cell(row=row_idx, column=7, value=g2 if found2 else None)
        result_ws.cell(row=row_idx, column=8, value=diff)
        result_ws.cell(row=row_idx, column=9, value=status)
        row_idx += 1

    last_data_row = row_idx - 1
    total_row = row_idx
    result_ws.cell(row=total_row, column=4, value="Grand Total").font = Font(bold=True)
    if last_data_row >= 2:
        result_ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{last_data_row})")
        result_ws.cell(row=total_row, column=7, value=f"=SUM(G2:G{last_data_row})")
        result_ws.cell(row=total_row, column=8, value=f"=SUM(H2:H{last_data_row})")
    for col in range(4, 10):
        cell = result_ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for col_idx in range(1, 10):
        result_ws.column_dimensions[get_column_letter(col_idx)].width = 18

    wb2.save(input_path2)
    return input_path2


def gross_salary_comparision(input_path, input_path2):
    """Converted from VBA: CompareGrossSalary (CommandButton13)."""
    return _compare_salary_field(input_path, input_path2, "Gross Salary")


def net_salary_comparision(input_path, input_path2):
    """Converted from VBA: CompareNetSalary (CommandButton15)."""
    return _compare_salary_field(input_path, input_path2, "Net Salary")


def _find_first_rate_column(ws, header_row, last_col):
    for c in range(1, last_col + 1):
        name = str(ws.cell(row=header_row, column=c).value or "")
        if name.replace(" ", "").strip().lower() == "rate":
            return c
    return 0


def tds_section_wise(input_path):
    """
    Converted from VBA: Cmd_TDS1_Click.
    Groups by Company > Section Code > Business Unit (each gets its own
    subtotal row, matching the VBA's default subtotal-on behavior for those
    3 fields), with Supplier as the detail row and no subtotal (matching
    the VBA explicitly turning Supplier's subtotal off). Headers fixed on
    row 15.
    """
    HEADER_ROW = 15

    wb = load_workbook(input_path, data_only=True)
    if len(wb.sheetnames) > 1:
        raise ValueError("Incorrect file. Please load the correct single-sheet file.")
    ws = wb.worksheets[0]

    last_col = 1
    for c in range(ws.max_column, 0, -1):
        if ws.cell(row=HEADER_ROW, column=c).value not in (None, ""):
            last_col = c
            break
    last_row = HEADER_ROW
    for r in range(ws.max_row, HEADER_ROW, -1):
        if ws.cell(row=r, column=1).value not in (None, ""):
            last_row = r
            break
    if last_row <= HEADER_ROW:
        raise ValueError("No data found below the headers.")

    headers = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, last_col + 1)]
    rows = [[ws.cell(row=r, column=c).value for c in range(1, last_col + 1)]
            for r in range(HEADER_ROW + 1, last_row + 1)]
    df = pd.DataFrame(rows, columns=headers)

    group_cols = ["Company", "Section Code", "Business Unit", "Supplier"]
    value_cols = ["Assessable Amount", "Net TDS Amount"]

    # Robust match (case/whitespace-insensitive) -- fixes false "missing
    # column" errors when the real file's header casing/spacing differs
    # slightly from what's expected.
    col_map = match_columns_case_insensitive(df, group_cols + value_cols)
    df = df.rename(columns={actual: wanted for wanted, actual in col_map.items()})

    for c in value_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    df = df.sort_values(by=group_cols).reset_index(drop=True)

    new_sheet_name = "Pivot_" + now_ist().strftime("%H%M%S")
    ws_out = wb.create_sheet(new_sheet_name)

    all_columns = group_cols + [f"Sum of {c}" for c in value_cols]
    df_for_write = df.rename(columns={c: f"Sum of {c}" for c in value_cols})
    write_hierarchical_table(
        ws_out, df_for_write,
        group_levels=["Company", "Section Code", "Business Unit"],
        all_columns=all_columns,
        value_cols=[f"Sum of {c}" for c in value_cols],
        start_row=1,
    )

    for col_idx in range(1, len(all_columns) + 1):
        ws_out.column_dimensions[get_column_letter(col_idx)].width = 18

    wb.save(input_path)
    return input_path


def tds_pivot_bu_wise(input_path):
    """
    Converted from VBA: Cmd_TDS2_Click.
    Groups by Business unit > Nature (each gets its own subtotal row,
    matching the VBA's default subtotal-on behavior for those 2 fields),
    with Supplier / Document Date / Rate as detail-only columns (matching
    the VBA explicitly turning their subtotals off). Headers fixed on
    row 15.
    """
    HEADER_ROW = 15

    wb = load_workbook(input_path, data_only=True)
    if len(wb.sheetnames) > 1:
        raise ValueError("Incorrect file. Please load the correct single-sheet file.")
    ws = wb.worksheets[0]

    last_col = 1
    for c in range(ws.max_column, 0, -1):
        if ws.cell(row=HEADER_ROW, column=c).value not in (None, ""):
            last_col = c
            break
    last_row = HEADER_ROW
    for r in range(ws.max_row, HEADER_ROW, -1):
        if ws.cell(row=r, column=1).value not in (None, ""):
            last_row = r
            break
    if last_row <= HEADER_ROW:
        raise ValueError("No data found below the headers.")

    rate_col_idx = _find_first_rate_column(ws, HEADER_ROW, last_col)

    headers = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, last_col + 1)]
    rows = [[ws.cell(row=r, column=c).value for c in range(1, last_col + 1)]
            for r in range(HEADER_ROW + 1, last_row + 1)]
    df = pd.DataFrame(rows, columns=headers)

    group_cols = ["Business unit", "Nature", "Supplier", "Document Date"]
    value_cols = ["Assessable Amount", "Net TDS Amount"]

    col_map = match_columns_case_insensitive(df, group_cols + value_cols)
    df = df.rename(columns={actual: wanted for wanted, actual in col_map.items()})

    rate_col_name = None
    if rate_col_idx:
        rate_col_name = headers[rate_col_idx - 1]
        if rate_col_name not in group_cols:
            group_cols = group_cols + [rate_col_name]

    for c in value_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    df["Document Date"] = pd.to_datetime(df["Document Date"], errors="coerce")
    df = df.sort_values(by=group_cols).reset_index(drop=True)

    new_sheet_name = "Pivot_" + now_ist().strftime("%H%M%S")
    ws_out = wb.create_sheet(new_sheet_name)

    all_columns = group_cols + [f"Sum of {c}" for c in value_cols]
    df_for_write = df.copy()
    df_for_write["Document Date"] = df_for_write["Document Date"].apply(
        lambda v: v.to_pydatetime() if pd.notna(v) else None
    )
    df_for_write = df_for_write.rename(columns={c: f"Sum of {c}" for c in value_cols})

    start_row = 1
    next_row = write_hierarchical_table(
        ws_out, df_for_write,
        group_levels=["Business unit", "Nature"],
        all_columns=all_columns,
        value_cols=[f"Sum of {c}" for c in value_cols],
        start_row=start_row,
    )

    # Apply number formats to the Document Date / Rate detail columns
    doc_date_col = all_columns.index("Document Date") + 1
    for r in range(start_row + 1, next_row):
        cell = ws_out.cell(row=r, column=doc_date_col)
        if cell.value is not None:
            cell.number_format = "DD/MM/YYYY"

    if rate_col_name:
        rate_col_idx_out = all_columns.index(rate_col_name) + 1
        for r in range(start_row + 1, next_row):
            cell = ws_out.cell(row=r, column=rate_col_idx_out)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    for col_idx in range(1, len(all_columns) + 1):
        ws_out.column_dimensions[get_column_letter(col_idx)].width = 18

    wb.save(input_path)
    return input_path


def gst_pivot_summary(input_path):
    # Skipped for now, per your instruction -- send the ProcessGSTReport
    # VBA code whenever you're ready and this will be wired up the same way.
    return {"message": "GST Pivot Summary isn't wired up yet."}


def sheet_exists(name, wb):
    """Equivalent of the VBA SheetExists helper."""
    return name.lower() in [s.lower() for s in wb.sheetnames]


def amount_received(input_path):
    """
    Converted from VBA: CreatePivotFromSourceData_Adv

    Builds a static "pivot-style" summary table (openpyxl/pandas cannot
    create a native, refreshable Excel PivotTable object -- this
    reproduces the same grouping/sums/formatting as a plain table instead).
    """
    wb = load_workbook(input_path)

    # --- Find SourceData sheet -------------------------------------------------
    if "SourceData" not in wb.sheetnames:
        raise ValueError("Sheet 'SourceData' not found in the selected file.")
    ws_data = wb["SourceData"]

    # --- Dynamic range: header row fixed at 13, last row/col detected ----------
    header_row = 13

    last_row = header_row
    for r in range(ws_data.max_row, header_row - 1, -1):
        if ws_data.cell(row=r, column=1).value not in (None, ""):
            last_row = r
            break

    last_col = 1
    for c in range(ws_data.max_column, 0, -1):
        if ws_data.cell(row=header_row, column=c).value not in (None, ""):
            last_col = c
            break

    # --- Read header + data rows into a DataFrame -------------------------------
    raw_headers = [ws_data.cell(row=header_row, column=c).value for c in range(1, last_col + 1)]
    # Some source files repeat the exact same header text for more than one
    # column (e.g. "TOTAL REC TILL DATE BASIC" appearing 2-3 times). Excel's
    # own PivotTable engine auto-renames repeats to "...2", "...3" internally
    # -- this reproduces that so those columns become individually usable.
    headers = dedupe_headers(raw_headers)

    data_rows = []
    for r in range(header_row + 1, last_row + 1):
        data_rows.append([ws_data.cell(row=r, column=c).value for c in range(1, last_col + 1)])

    df = pd.DataFrame(data_rows, columns=headers)

    required_cols = [
        "Status",
        "TOTAL REC TILL DATE BASIC",
        "On Account Amount",
        "TOTAL REC TILL DATE BASIC2",
        "On Account Amount2",
    ]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(
            "Missing expected column(s) in SourceData: " + ", ".join(missing)
        )

    value_cols = [
        "TOTAL REC TILL DATE BASIC",
        "On Account Amount",
        "TOTAL REC TILL DATE BASIC2",
        "On Account Amount2",
    ]
    for col in value_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # --- Group by Status, sum the value columns (the actual "pivot") -----------
    pivot = df.groupby("Status", dropna=False)[value_cols].sum()

    # --- New sheet name: Sheet1, Sheet2, Sheet3 ... -----------------------------
    i = 1
    while sheet_exists(f"Sheet{i}", wb):
        i += 1
    new_sheet_name = f"Sheet{i}"
    ws_pivot = wb.create_sheet(new_sheet_name)

    # --- Write the summary table starting at A3 (like TableDestination A3) -----
    start_row = 3
    start_col = 1  # column A

    header_labels = ["Row Labels"] + [f"Sum of {c}" for c in value_cols]
    for j, label in enumerate(header_labels):
        ws_pivot.cell(row=start_row, column=start_col + j, value=label)

    data_start_row = start_row + 1
    row_idx = data_start_row
    for status_val, row_data in pivot.iterrows():
        ws_pivot.cell(row=row_idx, column=start_col, value=status_val)
        for j, col in enumerate(value_cols):
            ws_pivot.cell(row=row_idx, column=start_col + 1 + j, value=float(row_data[col]))
        row_idx += 1

    grand_total_row = row_idx
    ws_pivot.cell(row=grand_total_row, column=start_col, value="Grand Total")
    for j, col in enumerate(value_cols):
        col_letter = get_column_letter(start_col + 1 + j)
        ws_pivot.cell(
            row=grand_total_row,
            column=start_col + 1 + j,
            value=f"=SUM({col_letter}{data_start_row}:{col_letter}{grand_total_row - 1})",
        )

    # --- G Total column (row-wise sum), like F3/F4/F5/F6 in the VBA ------------
    g_total_col = start_col + 1 + len(value_cols)  # column F when there are 4 value cols
    g_total_col_letter = get_column_letter(g_total_col)
    ws_pivot.cell(row=start_row, column=g_total_col, value="G Total")

    first_value_col_letter = get_column_letter(start_col + 1)
    last_value_col_letter = get_column_letter(start_col + len(value_cols))

    for r in range(data_start_row, grand_total_row):
        ws_pivot.cell(
            row=r,
            column=g_total_col,
            value=f"=SUM({first_value_col_letter}{r}:{last_value_col_letter}{r})",
        )

    ws_pivot.cell(
        row=grand_total_row,
        column=g_total_col,
        value=f"=SUM({g_total_col_letter}{data_start_row}:{g_total_col_letter}{grand_total_row - 1})",
    )

    # --- Borders + fill on the G Total column (matches the VBA formatting) -----
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

    for r in range(start_row, grand_total_row + 1):
        cell = ws_pivot.cell(row=r, column=g_total_col)
        cell.border = border
        cell.fill = fill

    # --- Save back into the same workbook (like wbSource.Save) ------------------
    wb.save(input_path)

    return input_path


def _normalize_value(v, col_type):
    """Equivalent of VBA NormalizeValue: makes values comparable regardless
    of minor type/formatting differences between the two files."""
    if col_type == "N":
        if v is None or str(v).strip() == "":
            return "0.00"
        try:
            return f"{float(v):.2f}"
        except (TypeError, ValueError):
            return str(v).strip().upper()
    elif col_type == "D":
        if v is None or str(v).strip() == "":
            return ""
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d")
        try:
            return pd.to_datetime(v).strftime("%Y-%m-%d")
        except Exception:
            return str(v).strip().upper()
    else:
        s = str(v).strip() if v is not None else ""
        while "  " in s:
            s = s.replace("  ", " ")
        return s.upper()


def daybook_difference(input_path, input_path2):
    """
    Converted from VBA: CompareToeWiseAbstract (CommandButton8, behind
    "1. Diffrence of Day Book A-B").

    Matches rows between the two files' "ToeWiseAbstract" sheets on
    (Document No, Debit, Credit) and reports rows that only exist in one
    file, or whose other columns differ.

    NOTE: the VBA also highlights matching/differing rows in green inside
    the SOURCE files themselves. This web version only writes the
    differences to the result sheet -- it doesn't modify the uploaded
    source files' own formatting, to keep from silently altering files
    beyond what gets downloaded.
    """
    COL_BUSINESS_UNIT = 3
    COL_DOC_DATE = 4
    COL_DOC_NO = 5
    COL_ACC_HEAD = 6
    COL_DEBIT = 7
    COL_CREDIT = 8
    COL_PARTY_LEDGER = 9
    HEADER_ROW = 1
    SHEET_NAME = "ToeWiseAbstract"

    compare_cols = [COL_BUSINESS_UNIT, COL_DOC_DATE, COL_DOC_NO, COL_ACC_HEAD, COL_DEBIT, COL_CREDIT, COL_PARTY_LEDGER]
    col_types = ["T", "D", "N", "T", "N", "N", "T"]
    out_headers = ["Business Unit", "Document Date", "Document No", "Account Head", "Debit", "Credit", "Party Ledger"]

    wb_a = load_workbook(input_path, data_only=True)
    wb_b = load_workbook(input_path2, data_only=True)

    if SHEET_NAME not in wb_a.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in File 1.")
    if SHEET_NAME not in wb_b.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in File 2.")

    ws_a = wb_a[SHEET_NAME]
    ws_b = wb_b[SHEET_NAME]

    def last_row_of(ws):
        last = ws.max_row
        for r in range(ws.max_row, HEADER_ROW, -1):
            if ws.cell(row=r, column=2).value not in (None, ""):
                return r
        return HEADER_ROW

    last_row_a = last_row_of(ws_a)
    last_row_b = last_row_of(ws_b)

    def row_key(ws, r):
        doc_no = _normalize_value(ws.cell(row=r, column=COL_DOC_NO).value, "N")
        debit = _normalize_value(ws.cell(row=r, column=COL_DEBIT).value, "N")
        credit = _normalize_value(ws.cell(row=r, column=COL_CREDIT).value, "N")
        return doc_no, debit, credit

    def row_values(ws, r):
        return [ws.cell(row=r, column=c).value for c in compare_cols]

    dict_b = {}
    for r in range(HEADER_ROW + 1, last_row_b + 1):
        doc_no, debit, credit = row_key(ws_b, r)
        if doc_no == "" and debit == "0.00" and credit == "0.00":
            continue
        dict_b.setdefault((doc_no, debit, credit), []).append(r)

    n = 1
    while f"Sheet{n}" in wb_b.sheetnames:
        n += 1
    new_sheet_name = f"Sheet{n}"
    ws_out = wb_b.create_sheet(new_sheet_name)

    ws_out.cell(row=1, column=1, value="Status")
    for c, h in enumerate(out_headers, start=2):
        ws_out.cell(row=1, column=c, value=h)

    out_row = 2

    def write_out_row(status, values):
        nonlocal out_row
        ws_out.cell(row=out_row, column=1, value=status)
        for c, v in enumerate(values, start=2):
            ws_out.cell(row=out_row, column=c, value=v)
        out_row += 1

    for r in range(HEADER_ROW + 1, last_row_a + 1):
        doc_no, debit, credit = row_key(ws_a, r)
        if doc_no == "" and debit == "0.00" and credit == "0.00":
            continue
        key = (doc_no, debit, credit)
        match_row = None
        if key in dict_b and dict_b[key]:
            match_row = dict_b[key].pop(0)

        if match_row is None:
            write_out_row("Exist in Sheet1", row_values(ws_a, r))
        else:
            is_diff = False
            for c, t in zip(compare_cols, col_types):
                va = _normalize_value(ws_a.cell(row=r, column=c).value, t)
                vb = _normalize_value(ws_b.cell(row=match_row, column=c).value, t)
                if va != vb:
                    is_diff = True
                    break
            if is_diff:
                write_out_row("Exist in Sheet1", row_values(ws_a, r))
                write_out_row("Exist in Sheet2", row_values(ws_b, match_row))

    for rows_left in dict_b.values():
        for r in rows_left:
            write_out_row("Exist in Sheet2", row_values(ws_b, r))

    for c in range(1, len(out_headers) + 2):
        ws_out.cell(row=1, column=c).font = Font(bold=True)
        ws_out.column_dimensions[get_column_letter(c)].width = 18

    wb_a.save(input_path)
    wb_b.save(input_path2)

    return input_path2


def creditor_outstanding(input_path):
    """
    Converted from VBA: GenerateCreditorReport

    Builds a two-section report on a new sheet:
      Section 1: "Outstanding Bills Amount" -> rows where Bill Outstanding <> 0
      Section 2: "Advance to Creditors"      -> rows where On A/C <> 0
    Each section gets a title row, a bold/shaded header row, its filtered
    data rows with a light grid, and a bold/shaded Total row.
    """

    # --- Same constants as the top of the VBA module ---------------------------
    HEADER_ROW = 6
    HDR_NO_OF_BILLS = "No. Of Bills"
    HDR_BILL_OUTSTANDING = "Bill Outstanding"
    HDR_ON_AC = "On A/C"
    HDR_NET_OUTSTANDING = "Net Outstanding"
    GAP_ROWS = 2

    wb = load_workbook(input_path)

    if "SourceData" not in wb.sheetnames:
        raise ValueError("Sheet 'SourceData' was not found in the selected file.")
    src_ws = wb["SourceData"]

    # --- last column (from header row) / last row (from column A) --------------
    last_col = 1
    for c in range(src_ws.max_column, 0, -1):
        if src_ws.cell(row=HEADER_ROW, column=c).value not in (None, ""):
            last_col = c
            break

    last_row = HEADER_ROW
    for r in range(src_ws.max_row, 0, -1):
        if src_ws.cell(row=r, column=1).value not in (None, ""):
            last_row = r
            break

    if last_row <= HEADER_ROW:
        raise ValueError("No data found below the headers in SourceData.")

    # --- find the needed columns by header text ---------------------------------
    col_no_of_bills = find_header_column(src_ws, HEADER_ROW, last_col, HDR_NO_OF_BILLS)
    col_bill_outstanding = find_header_column(src_ws, HEADER_ROW, last_col, HDR_BILL_OUTSTANDING)
    col_on_ac = find_header_column(src_ws, HEADER_ROW, last_col, HDR_ON_AC)
    col_net_outstanding = find_header_column(src_ws, HEADER_ROW, last_col, HDR_NET_OUTSTANDING)

    if col_bill_outstanding == 0 or col_on_ac == 0 or col_net_outstanding == 0:
        raise ValueError(
            f"Could not find '{HDR_BILL_OUTSTANDING}', '{HDR_ON_AC}' or '{HDR_NET_OUTSTANDING}' "
            f"headers in row {HEADER_ROW}. Please check the header name constants."
        )

    # --- columns to copy: every column except "No. Of Bills" --------------------
    src_cols = build_column_list(last_col, col_no_of_bills)

    # --- new sheet, placed right after SourceData --------------------------------
    dest_ws = add_next_available_sheet(wb, "SourceData")

    # --- timestamp -----------------------------------------------------------------
    dest_ws.cell(row=1, column=2, value="Report Generated: " + now_ist().strftime("%d-%b-%Y %I:%M:%S %p"))
    dest_ws.cell(row=1, column=2).font = Font(bold=True, italic=True)

    section_start_row = 3  # leave row 2 blank, same as the VBA

    # --- Section 1: Outstanding Bills Amount ------------------------------------
    next_start_row = write_section(
        dest_ws, src_ws, HEADER_ROW, last_row, src_cols,
        col_bill_outstanding, "Outstanding Bills Amount", section_start_row,
        col_bill_outstanding, col_on_ac, col_net_outstanding,
    )

    # --- Section 2: Advance to Creditors ----------------------------------------
    next_start_row = write_section(
        dest_ws, src_ws, HEADER_ROW, last_row, src_cols,
        col_on_ac, "Advance to Creditors", next_start_row + GAP_ROWS,
        col_bill_outstanding, col_on_ac, col_net_outstanding,
    )

    # --- reasonable column widths (openpyxl has no true "AutoFit") -------------
    for col_idx in range(2, len(src_cols) + 2):
        dest_ws.column_dimensions[get_column_letter(col_idx)].width = 16

    wb.save(input_path)
    return input_path


def find_header_column(ws, header_row, last_col, header_text):
    """Equivalent of VBA FindHeaderColumn: case/whitespace-insensitive header match."""
    target = header_text.strip().lower()
    for c in range(1, last_col + 1):
        val = ws.cell(row=header_row, column=c).value
        if val is not None and str(val).strip().lower() == target:
            return c
    return 0


def build_column_list(last_col, exclude_col):
    """Equivalent of VBA BuildColumnList: all source columns except one."""
    return [c for c in range(1, last_col + 1) if c != exclude_col]


def position_in_list(col_list, src_col_idx):
    """Equivalent of VBA PositionInList: 1-based position of a source column in col_list."""
    for i, c in enumerate(col_list, start=1):
        if c == src_col_idx:
            return i
    return 0


def add_next_available_sheet(wb, after_sheet_name):
    """Equivalent of VBA AddNextAvailableSheet: first unused 'SheetN' name, placed after a given sheet."""
    n = 1
    while f"Sheet{n}" in wb.sheetnames:
        n += 1
    new_name = f"Sheet{n}"
    idx = wb.sheetnames.index(after_sheet_name) + 1
    return wb.create_sheet(new_name, idx)


def write_section(dest_ws, src_ws, header_row, last_row, col_list, filter_src_col,
                   title_text, start_row, src_col_bill_outstanding, src_col_on_ac,
                   src_col_net_outstanding):
    """Equivalent of VBA WriteSection: title row + header row + filtered data rows + total row."""

    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    medium = Side(style="medium")

    title_fill = PatternFill(start_color="FBE0CE", end_color="FBE0CE", fill_type="solid")   # ~accent2, tint .6
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # ~accent1, tint .6
    total_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")   # ~accent1, tint .8

    last_dest_col = len(col_list)
    end_col = 1 + last_dest_col  # data occupies columns B..end_col

    pos_bill = position_in_list(col_list, src_col_bill_outstanding)
    pos_onac = position_in_list(col_list, src_col_on_ac)
    pos_net = position_in_list(col_list, src_col_net_outstanding)

    # --- Title row (merged, centered, shaded) -----------------------------------
    dest_ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=end_col)
    title_cell = dest_ws.cell(row=start_row, column=2, value=title_text)
    title_cell.alignment = Alignment(horizontal="center", vertical="bottom")
    for col in range(2, end_col + 1):
        cell = dest_ws.cell(row=start_row, column=col)
        cell.fill = title_fill
        cell.border = thin_border

    # --- Header row ---------------------------------------------------------------
    header_dest_row = start_row + 1
    dest_col = 2
    for src_c in col_list:
        dest_ws.cell(row=header_dest_row, column=dest_col, value=src_ws.cell(row=header_row, column=src_c).value)
        dest_col += 1
    for col in range(2, end_col + 1):
        cell = dest_ws.cell(row=header_dest_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # --- Data rows: only where filter_src_col value is numeric and non-zero ----
    dest_row = start_row + 2
    first_data_row = dest_row
    for r in range(header_row + 1, last_row + 1):
        val = src_ws.cell(row=r, column=filter_src_col).value
        if isinstance(val, (int, float)) and val != 0:
            dest_col = 2
            for src_c in col_list:
                dest_ws.cell(row=dest_row, column=dest_col, value=src_ws.cell(row=r, column=src_c).value)
                dest_col += 1
            dest_row += 1
    last_data_row = dest_row - 1

    if last_data_row >= first_data_row:
        for r in range(first_data_row, last_data_row + 1):
            for col in range(2, end_col + 1):
                dest_ws.cell(row=r, column=col).border = thin_border

    # --- Total row -----------------------------------------------------------------
    if last_data_row >= first_data_row:
        total_label_cell = dest_ws.cell(row=dest_row, column=2, value="Total")
        total_label_cell.font = Font(bold=True)

        for pos in (pos_bill, pos_onac, pos_net):
            if pos > 0:
                col_num = pos + 1
                letter = get_column_letter(col_num)
                cell = dest_ws.cell(row=dest_row, column=col_num)
                cell.value = f"=SUM({letter}{first_data_row}:{letter}{last_data_row})"
                cell.font = Font(bold=True)

        for col in range(2, end_col + 1):
            cell = dest_ws.cell(row=dest_row, column=col)
            cell.fill = total_fill
            cell.border = Border(left=thin, right=thin, top=medium, bottom=thin)

    return dest_row


# Maps the report_id sent from the HTML page to the function that handles it.
REPORT_HANDLERS = {
    "salary_pivot_bu_wise": salary_pivot_bu_wise,
    "salary_pivot_bu_wise_pdf": salary_pivot_bu_wise_pdf,
    "gross_salary_comparision": gross_salary_comparision,
    "net_salary_comparision": net_salary_comparision,
    "tds_section_wise": tds_section_wise,
    "tds_pivot_bu_wise": tds_pivot_bu_wise,
    "gst_pivot_summary": gst_pivot_summary,
    "amount_received": amount_received,
    "daybook_difference": daybook_difference,
    "creditor_outstanding": creditor_outstanding,
}

# Reports that compare TWO uploaded files (their handler takes
# (input_path, input_path2) instead of just (input_path)).
TWO_FILE_REPORTS = {
    "gross_salary_comparision",
    "net_salary_comparision",
    "daybook_difference",
}


@app.route("/")
def serve_login():
    return send_from_directory(".", "login_form.html")


@app.route("/dashboard.html")
def serve_dashboard():
    return send_from_directory(".", "dashboard.html")


@app.route("/api/run-report", methods=["POST"])
def run_report():
    try:
        report_id = request.form.get("report_id")
        uploaded_file = request.files.get("file")
        uploaded_file2 = request.files.get("file2")

        if not report_id or report_id not in REPORT_HANDLERS:
            return jsonify({"message": f"Unknown report_id: {report_id}"}), 400

        if not uploaded_file:
            return jsonify({"message": "No file was uploaded."}), 400

        if report_id in TWO_FILE_REPORTS and not uploaded_file2:
            return jsonify({"message": "This report needs a second file."}), 400

        # Each request gets its own private folder, and File 1 / File 2 each
        # get their OWN subfolder within it -- so even if both files happen
        # to share the exact same filename (e.g. both named "Report.xlsx"),
        # they can never overwrite each other.
        request_dir = os.path.join(BASE_TMP_DIR, uuid.uuid4().hex)
        file1_dir = os.path.join(request_dir, "file1")
        os.makedirs(file1_dir, exist_ok=True)

        input_path = os.path.join(file1_dir, uploaded_file.filename)
        uploaded_file.save(input_path)

        input_path2 = None
        if uploaded_file2:
            file2_dir = os.path.join(request_dir, "file2")
            os.makedirs(file2_dir, exist_ok=True)
            input_path2 = os.path.join(file2_dir, uploaded_file2.filename)
            uploaded_file2.save(input_path2)

        handler = REPORT_HANDLERS[report_id]

        if report_id in TWO_FILE_REPORTS:
            result = handler(input_path, input_path2)
        else:
            result = handler(input_path)

        # If the handler returned a file path, send that file back for download.
        if isinstance(result, str) and os.path.isfile(result):
            return send_file(result, as_attachment=True)

        # Otherwise treat it as a JSON status message.
        return jsonify(result)

    except Exception as exc:
        # Catches EVERY failure point in this route (file saving, report
        # logic, anything) so the browser always gets a clear JSON error
        # message instead of a raw "Internal Server Error" page.
        return jsonify({"message": f"Error while running the report: {exc}"}), 500


@app.errorhandler(Exception)
def handle_any_uncaught_error(exc):
    """Final safety net: if anything anywhere in the app raises an
    exception that wasn't already caught, return clean JSON instead of
    Flask's default HTML error page."""
    return jsonify({"message": f"Unexpected server error: {exc}"}), 500


if __name__ == "__main__":
    # For local testing only. When deployed, a production server
    # (gunicorn) runs this app instead -- see deployment notes.
    app.run(host="127.0.0.1", port=5000, debug=True)
